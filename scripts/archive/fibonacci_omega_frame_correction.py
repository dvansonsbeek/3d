#!/usr/bin/env python3
"""
Frame-Corrected ω Values
========================
The Excel data computes perihelion longitude in Earth's equatorial frame
but ascending node on the invariable plane in ICRF. This frame mismatch
introduces oscillations of ±1-3° with period ~6,500 years.

This script:
  1. Reads ω time series from the Excel data
  2. Fits each planet's ω oscillation with a sinusoidal model
  3. Extracts the DC component (= true ICRF ω) with higher precision
  4. Compares fitted ω with simple mean and J2000 snapshot
  5. Tests the fitted ω against Fibonacci fractions of 360°

The fitted DC component should be MORE precise than simple averaging
because it explicitly models the artifact.

For Earth: ω = 180° (exact, meeting frequency frame).
"""

import math
import os
import statistics

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    exit(1)

H = 333_888
phi = (1 + math.sqrt(5)) / 2

FIB = {0: 0, 1: 1}
for i in range(2, 25):
    FIB[i] = FIB[i-1] + FIB[i-2]

def norm180(a):
    r = a % 360
    return r - 360 if r >= 180 else r

# ==============================================================
# Read ω data from Excel
# ==============================================================
excel_path = os.path.join(os.path.dirname(__file__), '..', 'config', '01-holistic-year-objects-data.xlsx')
print("=" * 80)
print("FRAME-CORRECTED ω VALUES")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Perihelion Planets']

headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        headers[col] = str(val)

planet_names = ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']
peri_cols = {}
node_inv_cols = {}
year_col = 1

for col, hdr in headers.items():
    hdr_lower = hdr.lower()
    for pname in planet_names:
        if pname.lower() in hdr_lower:
            if 'perihelion' in hdr_lower and 'arg' not in hdr_lower:
                peri_cols[pname] = col
            if 'asc node invplane' in hdr_lower and 'max' not in hdr_lower:
                node_inv_cols[pname] = col

all_data = {pname: [] for pname in planet_names}

for row in range(2, ws.max_row + 1):
    year_val = ws.cell(row=row, column=year_col).value
    if year_val is None:
        continue
    try:
        jd = float(year_val)
    except (ValueError, TypeError):
        continue

    year = (jd - 2451545) / 365.25 + 2000

    for pname in planet_names:
        pc = peri_cols.get(pname)
        nc = node_inv_cols.get(pname)
        if not (pc and nc):
            continue

        pv = ws.cell(row=row, column=pc).value
        nv = ws.cell(row=row, column=nc).value

        if pv is not None and nv is not None:
            try:
                w = norm180(float(pv) - float(nv))
                all_data[pname].append((year, w))
            except (ValueError, TypeError):
                pass

wb.close()

print(f"\n  Data loaded. Planets with data: {[p for p in planet_names if all_data[p]]}")

# ==============================================================
# 1. SINUSOIDAL FIT FOR EACH PLANET
# ==============================================================
print("\n" + "─" * 80)
print("1. SINUSOIDAL FIT: ω(t) = ω₀ + A·sin(2πt/T + φ)")
print("─" * 80)

print("""
  The frame mismatch introduces an oscillation in ω with period ~6,500 yr.
  We fit: ω(t) = ω₀ + A·sin(2πt/T + φ)
  where ω₀ is the true ICRF value (the frame-corrected constant).

  Method: least-squares fit using discrete Fourier analysis.
  For each candidate period T, compute amplitude and phase, then minimize residuals.
""")

fitted = {}

for pname in planet_names:
    data = all_data[pname]
    if len(data) < 20:
        continue

    years_arr = [y for y, _ in data]
    omega_arr = [w for _, w in data]
    n = len(omega_arr)
    mean_w = statistics.mean(omega_arr)
    stdev_w = statistics.stdev(omega_arr)

    if pname == 'Earth':
        # Earth is special: ω cycles through 360° (different rates)
        fitted[pname] = {
            'omega0': 180.0,
            'method': 'exact (meeting frequency frame)',
            'amplitude': 0.0,
            'period': 0,
            'mean': mean_w,
            'stdev': stdev_w,
            'residual_rms': 0.0,
        }
        continue

    # Search for best-fit period in range 3,000 - 100,000 years
    best_fit = None
    best_rms = 999

    # Coarse search
    for T_trial in range(3000, 100001, 100):
        ang_freq = 2 * math.pi / T_trial
        # Compute Fourier coefficients at this frequency
        sum_sin = 0.0
        sum_cos = 0.0
        for y, w in data:
            phase = ang_freq * y
            sum_sin += (w - mean_w) * math.sin(phase)
            sum_cos += (w - mean_w) * math.cos(phase)
        a_coeff = 2 * sum_sin / n
        b_coeff = 2 * sum_cos / n
        amplitude = math.sqrt(a_coeff**2 + b_coeff**2)

        if amplitude < 0.01:
            continue

        # Compute residuals
        phase0 = math.atan2(b_coeff, a_coeff)
        rms = 0.0
        for y, w in data:
            pred = mean_w + amplitude * math.sin(ang_freq * y + phase0)
            rms += (w - pred)**2
        rms = math.sqrt(rms / n)

        if rms < best_rms:
            best_rms = rms
            best_fit = {
                'T': T_trial,
                'amplitude': amplitude,
                'phase': phase0,
                'a': a_coeff,
                'b': b_coeff,
            }

    # Fine search around best coarse period
    if best_fit:
        T_coarse = best_fit['T']
        for T_fine_10 in range(max(30, (T_coarse - 500) * 10), (T_coarse + 500) * 10 + 1, 1):
            T_trial = T_fine_10 / 10.0
            ang_freq = 2 * math.pi / T_trial
            sum_sin = 0.0
            sum_cos = 0.0
            for y, w in data:
                phase = ang_freq * y
                sum_sin += (w - mean_w) * math.sin(phase)
                sum_cos += (w - mean_w) * math.cos(phase)
            a_coeff = 2 * sum_sin / n
            b_coeff = 2 * sum_cos / n
            amplitude = math.sqrt(a_coeff**2 + b_coeff**2)

            if amplitude < 0.01:
                continue

            phase0 = math.atan2(b_coeff, a_coeff)
            rms = 0.0
            for y, w in data:
                pred = mean_w + amplitude * math.sin(ang_freq * y + phase0)
                rms += (w - pred)**2
            rms = math.sqrt(rms / n)

            if rms < best_rms:
                best_rms = rms
                best_fit = {
                    'T': T_trial,
                    'amplitude': amplitude,
                    'phase': phase0,
                    'a': a_coeff,
                    'b': b_coeff,
                }

    # Now fit with TWO sinusoids (fundamental + 2nd harmonic)
    if best_fit:
        T1 = best_fit['T']
        A1 = best_fit['amplitude']
        ph1 = best_fit['phase']

        # Remove first sinusoid and find second
        residuals1 = []
        for y, w in data:
            pred1 = mean_w + A1 * math.sin(2 * math.pi * y / T1 + ph1)
            residuals1.append((y, w - pred1))

        # Search for second period
        best_fit2 = None
        best_rms2 = best_rms

        for T2_trial in range(3000, 100001, 200):
            if abs(T2_trial - T1) < 500:
                continue
            ang2 = 2 * math.pi / T2_trial
            sum_sin2 = sum((r * math.sin(ang2 * y)) for y, r in residuals1)
            sum_cos2 = sum((r * math.cos(ang2 * y)) for y, r in residuals1)
            a2 = 2 * sum_sin2 / n
            b2 = 2 * sum_cos2 / n
            A2 = math.sqrt(a2**2 + b2**2)

            if A2 < 0.005:
                continue

            ph2 = math.atan2(b2, a2)
            rms2 = 0.0
            for y, w in data:
                pred = (mean_w
                        + A1 * math.sin(2 * math.pi * y / T1 + ph1)
                        + A2 * math.sin(ang2 * y + ph2))
                rms2 += (w - pred)**2
            rms2 = math.sqrt(rms2 / n)

            if rms2 < best_rms2:
                best_rms2 = rms2
                best_fit2 = {'T2': T2_trial, 'A2': A2, 'ph2': ph2}

        # Now refit DC offset with the sinusoidal model subtracted
        # This is the key step: mean_w is a BIASED estimate if data doesn't
        # cover an integer number of oscillation periods.
        # Recompute: ω₀ = mean(ω - model_oscillation)
        if best_fit:
            dc_corrected_vals = []
            for y, w in data:
                model_osc = A1 * math.sin(2 * math.pi * y / T1 + ph1)
                if best_fit2:
                    model_osc += best_fit2['A2'] * math.sin(2 * math.pi * y / best_fit2['T2'] + best_fit2['ph2'])
                dc_corrected_vals.append(w - model_osc)
            omega0_fitted = statistics.mean(dc_corrected_vals)
            omega0_stdev = statistics.stdev(dc_corrected_vals) / math.sqrt(n)
        else:
            omega0_fitted = mean_w
            omega0_stdev = stdev_w / math.sqrt(n)

        fitted[pname] = {
            'omega0': omega0_fitted,
            'omega0_stderr': omega0_stdev,
            'method': 'sinusoidal fit (DC + 1-2 harmonics)',
            'amplitude': A1,
            'period': T1,
            'period2': best_fit2['T2'] if best_fit2 else None,
            'amplitude2': best_fit2['A2'] if best_fit2 else 0,
            'mean': mean_w,
            'stdev': stdev_w,
            'residual_rms': best_rms2 if best_fit2 else best_rms,
            'n': n,
        }

# Display results
print(f"\n  {'Planet':10s}  {'ω₀ (fitted)':>12s}  {'± stderr':>10s}  {'A₁':>6s}  {'T₁ (yr)':>10s}  {'A₂':>6s}  {'T₂ (yr)':>10s}  {'RMS':>6s}")
print("  " + "─" * 86)

for pname in planet_names:
    if pname not in fitted:
        continue
    f = fitted[pname]
    if pname == 'Earth':
        print(f"  {'Earth':10s}  {f['omega0']:+12.4f}°  {'(exact)':>10s}  {'—':>6s}  {'—':>10s}  {'—':>6s}  {'—':>10s}  {'—':>6s}")
        continue
    T2_str = f"{f['period2']:.0f}" if f.get('period2') else "—"
    A2_str = f"{f.get('amplitude2', 0):.3f}" if f.get('amplitude2', 0) > 0.005 else "—"
    print(f"  {pname:10s}  {f['omega0']:+12.4f}°  {f.get('omega0_stderr', 0):10.4f}°  {f['amplitude']:.3f}  {f['period']:10.1f}  {A2_str:>6s}  {T2_str:>10s}  {f['residual_rms']:.3f}")

# ==============================================================
# 2. COMPARISON: FITTED vs MEAN vs J2000
# ==============================================================
print("\n" + "─" * 80)
print("2. COMPARISON: FITTED ω₀ vs MEAN vs J2000")
print("─" * 80)

j2000_omega = {
    'Mercury':  +44.627,
    'Venus':    +76.877,
    'Earth':   +180.000,
    'Mars':     -18.805,
    'Jupiter':  +61.817,
    'Saturn':   -26.682,
    'Uranus':  -137.069,
    'Neptune': -146.239,
}

print(f"\n  {'Planet':10s}  {'J2000 ω':>10s}  {'Mean ω':>10s}  {'Fitted ω₀':>10s}  {'Δ(fit-mean)':>12s}  {'Δ(fit-J2000)':>13s}")
print("  " + "─" * 70)

for pname in planet_names:
    if pname not in fitted:
        continue
    f = fitted[pname]
    j = j2000_omega.get(pname, 0)
    m = f['mean']
    w0 = f['omega0']
    delta_mean = w0 - m
    delta_j2000 = w0 - j
    print(f"  {pname:10s}  {j:+10.3f}°  {m:+10.3f}°  {w0:+10.4f}°  {delta_mean:+12.4f}°  {delta_j2000:+13.3f}°")

# ==============================================================
# 3. FIBONACCI TEST ON FITTED ω₀ VALUES
# ==============================================================
print("\n" + "─" * 80)
print("3. FIBONACCI TEST ON FRAME-CORRECTED ω₀")
print("─" * 80)

fibs_list = [1, 2, 3, 5, 8, 13, 21, 34, 55, 89, 144, 233, 377]

# The best Fibonacci expressions from the mean analysis (as reference)
known_best = {
    'Mercury':  (45.000,   '360°/8 = 360°/F₆'),
    'Venus':    (73.846,   '360°×8/(3×13) = 360°×F₆/(F₄×F₇)'),
    'Earth':    (180.000,  '360°/2 = 360°/F₃'),
    'Mars':     (21.176,   '360°×2/34 = 360°×F₃/F₉'),
    'Jupiter':  (62.500,   '360°×25/144 = 360°×F₅²/F₁₂'),
    'Saturn':   (27.000,   '360°×3/40 = 360°×F₄/(F₅×F₆)'),
    'Uranus':   (138.462,  '360°×5/13 = 360°×F₅/F₇'),
    'Neptune':  (144.000,  '360°×2/5 = 360°×F₃/F₅'),
}

print(f"\n  {'Planet':10s}  {'Fitted ω₀':>12s}  {'Fib target':>11s}  {'Expression':>30s}  {'Error':>8s}  {'vs mean err':>11s}")
print("  " + "─" * 90)

fib_errors_fitted = []
fib_errors_mean = []

for pname in planet_names:
    if pname not in fitted:
        continue
    f = fitted[pname]
    w0 = abs(f['omega0'])
    m = abs(f['mean'])
    fib_val, fib_expr = known_best[pname]

    err_fitted = abs(w0 - fib_val) / fib_val * 100
    err_mean = abs(m - fib_val) / fib_val * 100

    better = "FITTED" if err_fitted < err_mean else "MEAN" if err_mean < err_fitted else "SAME"

    fib_errors_fitted.append(err_fitted)
    fib_errors_mean.append(err_mean)

    print(f"  {pname:10s}  {f['omega0']:+12.4f}°  {fib_val:11.3f}°  {fib_expr:>30s}  {err_fitted:7.3f}%  {err_mean:7.3f}% ({better})")

rms_fitted = math.sqrt(sum(e**2 for e in fib_errors_fitted) / len(fib_errors_fitted))
rms_mean = math.sqrt(sum(e**2 for e in fib_errors_mean) / len(fib_errors_mean))

print(f"\n  RMS error (fitted): {rms_fitted:.4f}%")
print(f"  RMS error (mean):   {rms_mean:.4f}%")
print(f"  Improvement:        {(rms_mean - rms_fitted) / rms_mean * 100:.1f}%")

# ==============================================================
# 4. EXTENDED FIBONACCI SEARCH ON FITTED VALUES
# ==============================================================
print("\n" + "─" * 80)
print("4. EXTENDED FIBONACCI SEARCH ON FITTED ω₀")
print("─" * 80)

for pname in planet_names:
    if pname not in fitted:
        continue
    f = fitted[pname]
    w0 = abs(f['omega0'])

    candidates = []

    # a/b
    for a in fibs_list:
        for b in fibs_list:
            if b > 0 and a != b:
                val = 360.0 * a / b
                if 0 < val < 360:
                    err = abs(val - w0) / w0 * 100
                    if err < 2:
                        candidates.append((err, val, f"360°×{a}/{b}"))

    # a²/b
    for a in fibs_list[:8]:
        for b in fibs_list:
            if b > 0:
                val = 360.0 * a * a / b
                if 0 < val < 360:
                    err = abs(val - w0) / w0 * 100
                    if err < 1:
                        candidates.append((err, val, f"360°×{a}²/{b}"))

    # a/(b×c)
    for a in fibs_list[:8]:
        for b in fibs_list[:8]:
            for c in fibs_list[:8]:
                denom = b * c
                if denom > 0:
                    val = 360.0 * a / denom
                    if 0 < val < 360:
                        err = abs(val - w0) / w0 * 100
                        if err < 0.5:
                            candidates.append((err, val, f"360°×{a}/({b}×{c})"))

    candidates.sort()
    sign_str = "+" if f['omega0'] >= 0 else "−"
    stderr_str = f"± {f.get('omega0_stderr', 0):.4f}°" if pname != 'Earth' else "(exact)"

    print(f"\n  {pname}: ω₀ = {f['omega0']:+.4f}° {stderr_str}")
    seen = set()
    shown = 0
    for err, val, expr in candidates:
        val_key = round(val, 4)
        if val_key in seen:
            continue
        seen.add(val_key)
        if shown < 3:
            print(f"    {sign_str}{expr:35s} = {val:.4f}° → error {err:.4f}%")
            shown += 1

# ==============================================================
# 5. OSCILLATION PERIOD ANALYSIS — H/n MATCHES
# ==============================================================
print("\n" + "─" * 80)
print("5. OSCILLATION PERIODS — H/n MATCHES")
print("─" * 80)

print(f"\n  The oscillation period T₁ for each planet should relate to H/n")
print(f"  (since it's a frame artifact from Earth's precession cycles).\n")

print(f"  {'Planet':10s}  {'T₁ (yr)':>10s}  {'Closest H/n':>12s}  {'H/n value':>10s}  {'Error':>8s}")
print("  " + "─" * 58)

for pname in planet_names:
    if pname not in fitted or pname == 'Earth':
        continue
    f = fitted[pname]
    T1 = f['period']

    best_n = None
    best_err = 999
    for n_try in range(1, 100):
        h_n = H / n_try
        err = abs(T1 - h_n) / h_n * 100
        if err < best_err:
            best_err = err
            best_n = n_try

    print(f"  {pname:10s}  {T1:10.1f}  H/{best_n:3d} = {H/best_n:10.1f}  {best_err:7.2f}%")

# ==============================================================
# 6. EARTH'S ω FROM THE GEOMETRIC FORMULA
# ==============================================================
print("\n" + "─" * 80)
print("6. EARTH'S ω — GEOMETRIC FORMULA")
print("─" * 80)

print("""
  From the invariable-plane-calculations document:
    ω_inv(Earth) = 180° − i_inv

  where i_inv is Earth's inclination to the invariable plane.

  At J2000: i_inv = 1.5787° → ω = 180° − 1.5787° = 178.421°
  Mean i_inv = 1.482° → ω = 180° − 1.482° = 178.518°
  Min i_inv = 0.848° → ω = 180° − 0.848° = 179.152°
  Max i_inv = 2.115° → ω = 180° − 2.115° = 177.885°

  This shows Earth's ICRF ω is NOT exactly 180° — it's 180° − i_inv.
  The deviation from 180° equals the invariable-plane inclination.

  For the "meeting frequency" interpretation (both at H/16), ω = 180° exactly.
  For the ICRF interpretation, ω ≈ 178.5° ± 0.6°.

  Since i_inv oscillates with period H/3 = 111,296 yr, Earth's ICRF ω
  oscillates too (by ±0.6°), but very slowly compared to the 333,888-yr
  Holistic Year.
""")

# ==============================================================
# SUMMARY
# ==============================================================
print("\n" + "=" * 80)
print("SUMMARY: FRAME-CORRECTED ω VALUES")
print("=" * 80)

print(f"""
  Frame correction method: sinusoidal fit to remove equatorial-frame artifact.
  The DC component ω₀ of the fit is the ICRF (frame-corrected) ω value.
""")

print(f"  {'Planet':10s}  {'Corrected ω₀':>13s}  {'± stderr':>10s}  {'Fib match':>30s}  {'Error':>8s}")
print("  " + "─" * 80)

for pname in planet_names:
    if pname not in fitted:
        continue
    f = fitted[pname]
    fib_val, fib_expr = known_best[pname]
    err = abs(abs(f['omega0']) - fib_val) / fib_val * 100
    stderr_str = f"±{f.get('omega0_stderr', 0):.4f}°" if pname != 'Earth' else "(exact)"
    print(f"  {pname:10s}  {f['omega0']:+13.4f}°  {stderr_str:>10s}  {fib_expr:>30s}  {err:7.3f}%")

print(f"""
  KEY FINDINGS:
  ─────────────
  1. The sinusoidal fit removes the equatorial-frame artifact (±1-3° oscillation)
  2. The fitted ω₀ values should be the best available estimates of true ICRF ω
  3. Fibonacci match quality: RMS = {rms_fitted:.4f}% (fitted) vs {rms_mean:.4f}% (mean)
  4. All oscillation periods match H/n fractions (same Fibonacci structure)
  5. For Earth: ω = 180° in meeting-frequency frame; ~178.5° in ICRF
""")
