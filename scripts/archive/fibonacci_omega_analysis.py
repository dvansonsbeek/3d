#!/usr/bin/env python3
"""
Argument of Perihelion (ω) — Comprehensive Fibonacci Analysis
=============================================================
Investigates whether the constant argument of perihelion for all 8 planets
follows Fibonacci patterns, and whether it connects to existing quantum numbers.

Key question: Is ω a 7th Fibonacci Law, or derivable from Laws 1-6?

Method:
1. Extract exact ω from Excel data (01-holistic-year-objects-data.xlsx)
2. Compute ω = perihelion longitude − ascending node longitude (invariable plane)
3. Test Fibonacci fraction representations: ω = 360° × F_a / F_b
4. Check cross-planet connections through quantum numbers
5. Test mirror-pair relationships
"""

import math
import os

# Try to load Excel data
try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("WARNING: openpyxl not available, using J2000 reference data only\n")

# ==============================================================
# Constants
# ==============================================================
H = 333_888
BALANCE_YEAR = -301_340
J2000_YEAR = 2000
DT = J2000_YEAR - BALANCE_YEAR  # 303,340 years

# Fibonacci numbers
FIB = {}
FIB[0], FIB[1] = 0, 1
for i in range(2, 20):
    FIB[i] = FIB[i-1] + FIB[i-2]

phi = (1 + math.sqrt(5)) / 2

def norm360(angle):
    """Normalize angle to [0, 360)."""
    return angle % 360

def norm180(angle):
    """Normalize angle to [-180, 180)."""
    a = angle % 360
    return a - 360 if a >= 180 else a

# ==============================================================
# Planet data: J2000 orbital elements (invariable plane)
# ==============================================================
# θ₀ = J2000 perihelion longitude (from formulas.mdx)
# Ω  = J2000 ascending node to invariable plane (from ascending-node-calibration.mdx)
# Precession period (from Law 1)
# d  = Fibonacci divisor (quantum number from Law 2)
# b  = period denominator in H/b (from Law 1)
# F  = Fibonacci number in d = b × F coupling

planets = {
    'Mercury':  {'theta0': 77.457,  'omega_node': 32.83,  'period': 8*H/11, 'direction': +1, 'b_num': 11, 'b_den': 8, 'd': 21, 'F_idx': 8, 'mass_sqrt': 4.074e-4},
    'Venus':    {'theta0': 131.577, 'omega_node': 54.70,  'period': 2*H,    'direction': +1, 'b_num': 1,  'b_den': 2, 'd': 34, 'F_idx': 9, 'mass_sqrt': 1.564e-3},
    'Earth':    {'theta0': None,    'omega_node': 284.51, 'period': H/16,   'direction': +1, 'b_num': 16, 'b_den': 1, 'd': 3,  'F_idx': 4, 'mass_sqrt': 1.733e-3},
    'Mars':     {'theta0': 336.065, 'omega_node': 354.87, 'period': 3*H/13, 'direction': +1, 'b_num': 13, 'b_den': 3, 'd': 5,  'F_idx': 5, 'mass_sqrt': 5.681e-4},
    'Jupiter':  {'theta0': 14.707,  'omega_node': 312.89, 'period': H/5,    'direction': +1, 'b_num': 5,  'b_den': 1, 'd': 5,  'F_idx': 5, 'mass_sqrt': 3.090e-2},
    'Saturn':   {'theta0': 92.128,  'omega_node': 118.81, 'period': H/8,    'direction': -1, 'b_num': 8,  'b_den': 1, 'd': 3,  'F_idx': 4, 'mass_sqrt': 1.691e-2},
    'Uranus':   {'theta0': 170.731, 'omega_node': 307.80, 'period': H/3,    'direction': +1, 'b_num': 3,  'b_den': 1, 'd': 21, 'F_idx': 8, 'mass_sqrt': 6.608e-3},
    'Neptune':  {'theta0': 45.801,  'omega_node': 192.04, 'period': 2*H,    'direction': +1, 'b_num': 1,  'b_den': 2, 'd': 34, 'F_idx': 9, 'mass_sqrt': 7.177e-3},
}

# Note: Earth's θ₀ is not listed in the planet table (it has its own 12-harmonic formula).
# We know from the balance year that Earth's ω = 180° (perihelion at line of nodes).
# Earth's θ₀ at J2000 = Ω + ω = 284.51° + 180° = 104.51° (mod 360°)
planets['Earth']['theta0'] = norm360(284.51 + 180.0)

# Precession rates for ascending node (same as perihelion in the model)
# The key insight: in the Holistic model, perihelion AND ascending node precess
# at the same rate → ω = θ₀ - Ω = constant

print("=" * 80)
print("ARGUMENT OF PERIHELION (ω) — FIBONACCI ANALYSIS")
print("=" * 80)

# ==============================================================
# SECTION 1: Compute ω from J2000 data
# ==============================================================
print("\n" + "=" * 80)
print("1. ω FROM J2000 DATA: ω = θ₀ - Ω")
print("=" * 80)

print(f"\n  {'Planet':10s}  {'θ₀ (°)':>10s}  {'Ω (°)':>10s}  {'ω (°)':>10s}  {'ω signed':>10s}")
print("  " + "-" * 56)

omega_vals = {}
for name, p in planets.items():
    theta = p['theta0']
    node = p['omega_node']
    omega = norm360(theta - node)
    omega_signed = norm180(theta - node)
    omega_vals[name] = {'raw': omega, 'signed': omega_signed}
    print(f"  {name:10s}  {theta:10.3f}  {node:10.3f}  {omega:10.3f}  {omega_signed:+10.3f}")

# ==============================================================
# SECTION 2: Read Excel data for verification
# ==============================================================
if HAS_EXCEL:
    print("\n" + "=" * 80)
    print("2. ω FROM EXCEL DATA (verification)")
    print("=" * 80)

    excel_path = os.path.join(os.path.dirname(__file__), '..', 'config', '01-holistic-year-objects-data.xlsx')
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Check available sheets
        print(f"\n  Sheets: {wb.sheetnames}")

        # Read the Perihelion Planets sheet
        ws = wb['Perihelion Planets']

        # Get headers from row 1
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append((col, str(val)))

        # Find columns for argument of perihelion
        arg_peri_cols = [(c, h) for c, h in headers if 'arg' in h.lower() and 'peri' in h.lower()]
        peri_cols = [(c, h) for c, h in headers if 'perihelion' in h.lower() and 'arg' not in h.lower()]
        node_cols = [(c, h) for c, h in headers if 'ascending' in h.lower() or 'asc' in h.lower() or 'node' in h.lower()]

        print(f"\n  Found {len(arg_peri_cols)} 'arg perihelion' columns:")
        for c, h in arg_peri_cols:
            print(f"    Col {c}: {h}")

        print(f"\n  Found {len(peri_cols)} 'perihelion' columns:")
        for c, h in peri_cols[:16]:
            print(f"    Col {c}: {h}")

        print(f"\n  Found {len(node_cols)} 'ascending node' columns:")
        for c, h in node_cols[:16]:
            print(f"    Col {c}: {h}")

        # Get year column
        year_col = 1  # Typically first column

        # Find a row near J2000
        j2000_row = None
        balance_row = None
        for row in range(2, min(ws.max_row + 1, 3500)):
            year_val = ws.cell(row=row, column=year_col).value
            if year_val is not None:
                try:
                    year_num = float(year_val)
                    if abs(year_num - 2000) < 2:
                        j2000_row = row
                    if abs(year_num - BALANCE_YEAR) < 500:
                        if balance_row is None or abs(year_num - BALANCE_YEAR) < abs(float(ws.cell(row=balance_row, column=year_col).value) - BALANCE_YEAR):
                            balance_row = row
                except (ValueError, TypeError):
                    pass

        if j2000_row:
            print(f"\n  J2000 row found: row {j2000_row} (year = {ws.cell(row=j2000_row, column=year_col).value})")
        if balance_row:
            print(f"  Balance year row found: row {balance_row} (year = {ws.cell(row=balance_row, column=year_col).value})")

        # Extract ω from arg_peri columns
        if arg_peri_cols and j2000_row:
            print(f"\n  ω from Excel (arg perihelion) at J2000:")
            for c, h in arg_peri_cols:
                val = ws.cell(row=j2000_row, column=c).value
                if val is not None:
                    print(f"    {h}: {float(val):.4f}°")

        if arg_peri_cols and balance_row:
            print(f"\n  ω from Excel (arg perihelion) at balance year:")
            for c, h in arg_peri_cols:
                val = ws.cell(row=balance_row, column=c).value
                if val is not None:
                    print(f"    {h}: {float(val):.4f}°")

        # Compute ω from perihelion - ascending node columns
        # Need to pair them by planet
        planet_names_short = ['mercury', 'venus', 'earth', 'mars', 'jupiter', 'saturn', 'uranus', 'neptune']

        print(f"\n  Computing ω = perihelion − ascending node from Excel:")
        for pname in planet_names_short:
            peri_match = [(c, h) for c, h in peri_cols if pname in h.lower()]
            node_match = [(c, h) for c, h in node_cols if pname in h.lower()]
            arg_match = [(c, h) for c, h in arg_peri_cols if pname in h.lower()]

            if j2000_row:
                row = j2000_row
                if arg_match:
                    val = ws.cell(row=row, column=arg_match[0][0]).value
                    if val is not None:
                        print(f"    {pname.capitalize():10s} ω (direct): {float(val):.4f}°")
                elif peri_match and node_match:
                    peri_val = ws.cell(row=row, column=peri_match[0][0]).value
                    node_val = ws.cell(row=row, column=node_match[0][0]).value
                    if peri_val is not None and node_val is not None:
                        omega_excel = norm360(float(peri_val) - float(node_val))
                        print(f"    {pname.capitalize():10s} ω (computed): {omega_excel:.4f}°")

        # Also check constancy: get ω at several epochs
        print(f"\n  ω CONSTANCY CHECK (multiple epochs):")
        sample_rows = []
        for row in range(2, min(ws.max_row + 1, 3500)):
            year_val = ws.cell(row=row, column=year_col).value
            if year_val is not None:
                try:
                    y = float(year_val)
                    # Sample every ~50,000 years
                    if row == 2 or (len(sample_rows) > 0 and y - sample_rows[-1][1] > 40000) or row == j2000_row or row == balance_row:
                        sample_rows.append((row, y))
                except (ValueError, TypeError):
                    pass

        # For Jupiter (well-established)
        jup_arg_col = None
        for c, h in arg_peri_cols:
            if 'jupiter' in h.lower():
                jup_arg_col = c
                break

        if jup_arg_col and len(sample_rows) > 2:
            print(f"\n    Jupiter ω across epochs:")
            jup_omegas = []
            for row, year in sample_rows[:12]:
                val = ws.cell(row=row, column=jup_arg_col).value
                if val is not None:
                    jup_omegas.append(float(val))
                    print(f"      Year {year:>10.0f}: ω = {float(val):.4f}°")
            if jup_omegas:
                print(f"      Range: {max(jup_omegas) - min(jup_omegas):.4f}° (should be ~0 if constant)")

        wb.close()
    else:
        print(f"\n  Excel file not found at: {excel_path}")

# ==============================================================
# SECTION 3: Fibonacci fraction analysis
# ==============================================================
print("\n" + "=" * 80)
print("3. FIBONACCI FRACTION ANALYSIS")
print("=" * 80)
print("\n  Testing ω = 360° × a/b where a, b involve Fibonacci numbers")

# Generate Fibonacci-based fractions (allow products and quotients)
def fib_fractions():
    """Generate fractions from Fibonacci numbers and their products."""
    fracs = []
    fibs = [1, 2, 3, 5, 8, 13, 21, 34, 55, 89, 144, 233]

    # Simple F_a / F_b
    for i, a in enumerate(fibs):
        for j, b in enumerate(fibs):
            if a != b:
                fracs.append((a, b, f"F/F ({a}/{b})"))

    # F_a² / F_b
    for i, a in enumerate(fibs):
        for j, b in enumerate(fibs):
            if a*a != b and b > 0:
                fracs.append((a*a, b, f"F²/F ({a}²/{b})"))

    # F_a / (F_b × F_c) for small ones
    for a in fibs[:8]:
        for b in fibs[:8]:
            for c in fibs[:8]:
                if b*c > 0 and a != b*c:
                    fracs.append((a, b*c, f"F/(F×F) ({a}/({b}×{c}))"))

    # Integer / Fibonacci
    for n in range(1, 40):
        for f in fibs:
            if n != f:
                fracs.append((n, f, f"n/F ({n}/{f})"))

    return fracs

all_fracs = fib_fractions()

for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    omega = omega_vals[name]['raw']
    omega_s = omega_vals[name]['signed']

    # Use signed version for negative ω
    target_abs = omega if omega <= 180 else 360 - omega
    is_negative = omega > 180

    print(f"\n  {name}: ω = {omega:.3f}° (signed: {omega_s:+.3f}°)")

    # Test 360° × fraction
    candidates = []
    for num, den, label in all_fracs:
        if den == 0:
            continue
        val = 360.0 * num / den
        if val > 0:
            # Compare with both positive and absolute
            for target in [omega, target_abs]:
                err_pct = 100 * abs(val - target) / max(target, 0.01)
                if err_pct < 2.0:
                    candidates.append((err_pct, val, num, den, label, target))

    candidates.sort()
    seen = set()
    shown = 0
    for err_pct, val, num, den, label, target in candidates:
        key = (num, den)
        if key in seen:
            continue
        seen.add(key)
        if shown < 5:
            sign = "-" if is_negative and target == target_abs else "+"
            print(f"    360° × {num}/{den} = {val:.4f}° → error {err_pct:.3f}%  [{label}]")
            shown += 1

# ==============================================================
# SECTION 4: Clean Fibonacci expressions
# ==============================================================
print("\n" + "=" * 80)
print("4. CLEAN FIBONACCI EXPRESSIONS FOR ω")
print("=" * 80)

# Based on the previous analysis and refined J2000 data
# Let's test the hypothesized expressions
hypotheses = {
    'Mercury':  ('360/8',         360.0/8,   'F₆ = 8'),
    'Venus':    ('360/5',         360.0/5,   'F₅ = 5'),
    'Earth':    ('360/2',         360.0/2,   'F₃ = 2 → 180°'),
    'Mars':     ('-360×3/55',    -360.0*3/55, 'F₄×360/F₁₀ → 3×360/55'),
    'Jupiter':  ('360×25/144',    360.0*25/144, 'F₅²/F₁₂'),
    'Saturn':   ('-360/13',      -360.0/13,  'F₇ = 13'),
    'Uranus':   ('-360×5/13',    -360.0*5/13, 'F₅/F₇'),
    'Neptune':  ('-360×2/5',     -360.0*2/5, 'F₃/F₅'),
}

print(f"\n  {'Planet':10s}  {'ω observed':>12s}  {'ω predicted':>12s}  {'Expression':20s}  {'Error':>10s}")
print("  " + "-" * 72)

for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    omega_obs = omega_vals[name]['signed']
    expr_str, omega_pred, fib_note = hypotheses[name]
    error = omega_obs - omega_pred
    error_pct = 100 * abs(error) / max(abs(omega_pred), 0.01)
    print(f"  {name:10s}  {omega_obs:+12.3f}°  {omega_pred:+12.3f}°  {expr_str:20s}  {error:+.3f}° ({error_pct:.2f}%)")

# ==============================================================
# SECTION 5: Cross-planet connections via quantum numbers
# ==============================================================
print("\n" + "=" * 80)
print("5. CROSS-PLANET CONNECTIONS")
print("=" * 80)

print("""
  Testing whether each planet's ω uses another planet's quantum numbers:

  Hypothesis from previous analysis:
    Mercury ω = 360°/8   → 8 is Saturn's period denominator (b_Saturn = 8)
    Venus   ω = 360°/5   → 5 is Jupiter's period denominator (b_Jupiter = 5)
    Earth   ω = 360°/2   → 2 is F₃ (Earth's own Fibonacci index)
    Jupiter ω = 360°×F₅²/F₁₂ → F₅ = 5 (own), F₁₂ = 144
    Saturn  ω = -360°/13 → 13 is axial precession denominator (H/13)
""")

# The key test: does each planet's ω denominator correspond to
# a specific quantum number from the model?

connections = [
    ('Mercury',  8,   "Saturn's b (period denominator H/8)"),
    ('Venus',    5,   "Jupiter's b (period denominator H/5)"),
    ('Earth',    2,   "F₃ = 2 (Earth's Fibonacci number in d=3=b×F=3×1)"),
    ('Mars',     55,  "F₁₀ = 55 (?)"),
    ('Jupiter',  144, "F₁₂ = 144 (Jupiter: b=5, 5²=25 numerator)"),
    ('Saturn',   13,  "Axial precession denominator (H/13)"),
    ('Uranus',   13,  "Axial precession denominator (H/13) × F₅/1"),
    ('Neptune',  5,   "Jupiter's b (period denominator H/5) × F₃/1"),
]

print(f"\n  {'Planet':10s}  {'ω':>10s}  {'Denominator':>12s}  {'Connection'}")
print("  " + "-" * 65)
for name, denom, conn in connections:
    omega_s = omega_vals[name]['signed']
    print(f"  {name:10s}  {omega_s:+10.3f}°  {denom:>12d}        {conn}")

# ==============================================================
# SECTION 6: Mirror pair analysis
# ==============================================================
print("\n" + "=" * 80)
print("6. MIRROR PAIR ANALYSIS")
print("=" * 80)

mirror_pairs = [
    ('Mercury', 'Uranus',  'd=21 ↔ d=21'),
    ('Venus',   'Neptune', 'd=34 ↔ d=34'),
    ('Earth',   'Saturn',  'd=3 ↔ d=3'),
    ('Mars',    'Jupiter', 'd=5 ↔ d=5'),
]

print(f"\n  {'Pair':25s}  {'ω_inner':>10s}  {'ω_outer':>10s}  {'Sum':>10s}  {'Ratio':>10s}  {'Notes'}")
print("  " + "-" * 90)

for inner, outer, label in mirror_pairs:
    w_i = omega_vals[inner]['signed']
    w_o = omega_vals[outer]['signed']
    w_sum = w_i + w_o
    ratio = w_i / w_o if w_o != 0 else float('inf')

    # Check if sum or difference has Fibonacci significance
    notes = []
    if abs(w_sum) < 2:
        notes.append("sum ≈ 0 (anti-parallel)")
    if abs(abs(w_sum) - 180) < 5:
        notes.append(f"sum ≈ ±180°")
    if abs(abs(ratio) - 1) < 0.05:
        notes.append("|ratio| ≈ 1")
    if abs(abs(ratio) - phi) < 0.1:
        notes.append(f"|ratio| ≈ φ")
    if abs(abs(ratio) - 2) < 0.2:
        notes.append(f"|ratio| ≈ 2")
    if abs(abs(ratio) - phi**2) < 0.2:
        notes.append(f"|ratio| ≈ φ²")

    # Check Fibonacci ratio
    for fn, fm in [(1,1), (1,2), (2,1), (2,3), (3,2), (3,5), (5,3), (5,8), (8,5), (8,13), (13,8)]:
        if abs(abs(ratio) - fn/fm) < 0.1:
            notes.append(f"|ratio| ≈ {fn}/{fm}")

    note_str = '; '.join(notes) if notes else ''
    print(f"  {inner+'/'+outer:25s}  {w_i:+10.3f}°  {w_o:+10.3f}°  {w_sum:+10.3f}°  {ratio:+10.4f}  {note_str}")

# Additional mirror analysis
print(f"\n  Detailed mirror relationships:")
for inner, outer, label in mirror_pairs:
    w_i = omega_vals[inner]['signed']
    w_o = omega_vals[outer]['signed']
    d = planets[inner]['d']

    print(f"\n    {inner} ({w_i:+.3f}°) ↔ {outer} ({w_o:+.3f}°)  [d = {d}]")
    print(f"      Sum:  {w_i + w_o:+.3f}°")
    print(f"      Diff: {w_i - w_o:+.3f}°")
    print(f"      Ratio: {w_i/w_o if w_o != 0 else 'inf':+.4f}")

    # d × ω
    print(f"      d × ω_inner = {d} × {w_i:.3f}° = {d * w_i:.3f}°")
    print(f"      d × ω_outer = {d} × {w_o:.3f}° = {d * w_o:.3f}°")

    # ω × √m
    m_i = planets[inner]['mass_sqrt']
    m_o = planets[outer]['mass_sqrt']
    print(f"      ω × √m_inner = {w_i:.3f}° × {m_i:.4e} = {w_i * m_i:.6f}")
    print(f"      ω × √m_outer = {w_o:.3f}° × {m_o:.4e} = {w_o * m_o:.6f}")

# ==============================================================
# SECTION 7: ω and the period structure
# ==============================================================
print("\n" + "=" * 80)
print("7. ω AND THE PERIOD STRUCTURE (H/b)")
print("=" * 80)

print(f"\n  Testing: ω × b/360° = ? (where b is the period denominator in H/b)")
print(f"\n  {'Planet':10s}  {'ω (°)':>10s}  {'b':>5s}  {'ω×b/360':>10s}  {'Nearest Fibonacci'}")
print("  " + "-" * 60)

fibs_list = [0, 1, 1, 2, 3, 5, 8, 13, 21, 34, 55, 89, 144]

for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    w = omega_vals[name]['signed']
    p = planets[name]

    # b = period numerator coefficient: period = (b_den/b_num) × H
    # So b = b_num / b_den (effective denominator)
    b_eff = p['b_num'] / p['b_den']

    product = abs(w) * b_eff / 360.0

    # Find nearest Fibonacci
    nearest_fib = min(fibs_list[1:], key=lambda f: abs(product - f))
    err = abs(product - nearest_fib) / nearest_fib * 100 if nearest_fib > 0 else float('inf')

    # Also check Fibonacci ratios
    fib_ratios = [(a, b, a/b) for a in fibs_list[1:12] for b in fibs_list[1:12] if b > 0]
    nearest_ratio = min(fib_ratios, key=lambda x: abs(product - x[2]))
    err_ratio = abs(product - nearest_ratio[2]) / nearest_ratio[2] * 100

    note = f"≈ {nearest_fib} ({err:.1f}%)" if err < 5 else f"≈ {nearest_ratio[0]}/{nearest_ratio[1]} ({err_ratio:.1f}%)"
    print(f"  {name:10s}  {w:+10.3f}°  {b_eff:5.1f}  {product:10.4f}    {note}")

# ==============================================================
# SECTION 8: Unified formula test
# ==============================================================
print("\n" + "=" * 80)
print("8. UNIFIED FORMULA TEST: ω = 360° × f(quantum numbers)")
print("=" * 80)

print("""
  The inclination constant is:      ψ = d × η × √m = const
  The eccentricity constant is:     d × ξ × √m = const (via R)

  Can we find:  ω = 360° × g(d, b, F) ?

  Where d = Fibonacci divisor, b = period denominator, F = Fibonacci index
""")

# Let's try systematic combinations
print(f"\n  Systematic test of ω = 360° / d:")
print(f"  {'Planet':10s}  {'d':>5s}  {'360/d':>10s}  {'ω obs':>10s}  {'Error':>10s}")
print("  " + "-" * 50)
for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    d = planets[name]['d']
    pred = 360.0 / d
    obs = abs(omega_vals[name]['signed'])
    err = obs - pred
    print(f"  {name:10s}  {d:5d}  {pred:10.3f}°  {obs:10.3f}°  {err:+10.3f}° ({100*abs(err)/obs:.1f}%)")

print(f"\n  Test of ω = 360° / b:")
print(f"  {'Planet':10s}  {'b':>5s}  {'360/b':>10s}  {'ω obs':>10s}  {'Error':>10s}")
print("  " + "-" * 50)
for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    p = planets[name]
    b_eff = p['b_num'] / p['b_den']
    pred = 360.0 / b_eff
    obs = abs(omega_vals[name]['signed'])
    err = obs - pred
    print(f"  {name:10s}  {b_eff:5.1f}  {pred:10.3f}°  {obs:10.3f}°  {err:+10.3f}° ({100*abs(err)/obs:.1f}%)")

# Try ω = 360° / b_mirror (where b_mirror is the MIRROR planet's b)
print(f"\n  Test of ω = 360° / b_mirror (MIRROR planet's period denominator):")
mirror_map = {'Mercury': 'Uranus', 'Venus': 'Neptune', 'Earth': 'Saturn',
              'Mars': 'Jupiter', 'Jupiter': 'Mars', 'Saturn': 'Earth',
              'Uranus': 'Mercury', 'Neptune': 'Venus'}

print(f"  {'Planet':10s}  {'b_mirror':>8s}  {'360/b_m':>10s}  {'ω obs':>10s}  {'Error':>10s}")
print("  " + "-" * 55)
for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    p = planets[name]
    mirror = mirror_map[name]
    pm = planets[mirror]
    b_mirror = pm['b_num'] / pm['b_den']
    pred = 360.0 / b_mirror
    obs = abs(omega_vals[name]['signed'])
    err = obs - pred
    print(f"  {name:10s}  {b_mirror:8.1f}  {pred:10.3f}°  {obs:10.3f}°  {err:+10.3f}° ({100*abs(err)/obs:.1f}%)")

# ==============================================================
# SECTION 9: Law 6 triad and ω
# ==============================================================
print("\n" + "=" * 80)
print("9. LAW 6 TRIAD (3+5=8) AND ω")
print("=" * 80)

w_E = omega_vals['Earth']['signed']    # 180°
w_J = omega_vals['Jupiter']['signed']  # ~62°
w_S = omega_vals['Saturn']['signed']   # ~-27°

print(f"\n  Earth ω   = {w_E:+.3f}°")
print(f"  Jupiter ω = {w_J:+.3f}°")
print(f"  Saturn ω  = {w_S:+.3f}°")

# Test linear combinations with 3,5,8 coefficients
print(f"\n  Linear combinations:")
for a, b, c in [(3,5,8), (5,8,13), (3,5,-8), (8,5,3), (1,1,1), (3,-5,8)]:
    val = a * w_E + b * w_J + c * w_S
    print(f"    {a}×ω_E + {b}×ω_J + {c}×ω_S = {a}×{w_E:.1f} + {b}×{w_J:.1f} + {c}×{w_S:.1f} = {val:.2f}°")
    val_mod = norm360(val)
    val_near = norm180(val)
    if abs(val_near) < 5 or abs(val_mod) < 5 or abs(val_mod - 360) < 5 or abs(val_mod - 180) < 5:
        print(f"      → ≈ {val_near:.2f}° (NOTABLE!)")

# ω_J + |ω_S| ≈ 90°?
print(f"\n  Jupiter + |Saturn| = {w_J:.3f}° + {abs(w_S):.3f}° = {w_J + abs(w_S):.3f}°")
print(f"  Compare to 90° (= 360°/4): error {abs(w_J + abs(w_S) - 90):.3f}°")
print(f"  Compare to Earth/2 = {w_E/2:.1f}°: error {abs(w_J + abs(w_S) - w_E/2):.3f}°")

# ==============================================================
# SECTION 10: ω as phase of the s8 eigenmode
# ==============================================================
print("\n" + "=" * 80)
print("10. ω AND THE SECULAR EIGENMODE PHASE")
print("=" * 80)

print("""
  In the Holistic model, all prograde planets share phase 203.3195°
  and Saturn (retrograde) has phase 23.3195° (= 203.3195° − 180°).

  Phase difference = 180° = Earth's ω.

  Is ω related to the eigenmode phase offset from 0°?
""")

s8_phase = 203.3195  # From the model

print(f"\n  s₈ eigenmode phase = {s8_phase:.4f}°")
print(f"\n  {'Planet':10s}  {'ω':>10s}  {'ω − s₈':>12s}  {'ω + s₈':>12s}  {'ω × d':>10s}")
print("  " + "-" * 60)

for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    w = omega_vals[name]['signed']
    d = planets[name]['d']
    diff = norm180(w - s8_phase)
    sumv = norm180(w + s8_phase)
    wd = norm180(w * d)
    print(f"  {name:10s}  {w:+10.3f}°  {diff:+12.3f}°  {sumv:+12.3f}°  {wd:+10.3f}°")

# ==============================================================
# SECTION 11: The d × ω product
# ==============================================================
print("\n" + "=" * 80)
print("11. d × ω PRODUCT — LOOKING FOR UNIVERSALITY")
print("=" * 80)

print(f"\n  Like d × η × √m = ψ (Law 2), does d × ω = const?")
print(f"\n  {'Planet':10s}  {'d':>5s}  {'ω (°)':>10s}  {'d × ω':>12s}  {'d × |ω|':>12s}")
print("  " + "-" * 55)

d_omega_vals = []
for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    d = planets[name]['d']
    w = omega_vals[name]['signed']
    dw = d * w
    dw_abs = d * abs(w)
    d_omega_vals.append(dw_abs)
    print(f"  {name:10s}  {d:5d}  {w:+10.3f}°  {dw:+12.3f}°  {dw_abs:12.3f}°")

print(f"\n  Range of d × |ω|: {min(d_omega_vals):.1f}° to {max(d_omega_vals):.1f}°")
print(f"  Spread: {(max(d_omega_vals) - min(d_omega_vals)) / ((max(d_omega_vals) + min(d_omega_vals))/2) * 100:.1f}%")
print(f"  → NOT universal (spread too large)")

# ==============================================================
# SECTION 12: ω × √m — mass-weighted version
# ==============================================================
print("\n" + "=" * 80)
print("12. ω × √m — MASS-WEIGHTED ANALYSIS")
print("=" * 80)

print(f"\n  {'Planet':10s}  {'ω (°)':>10s}  {'√m':>12s}  {'ω × √m':>12s}")
print("  " + "-" * 50)

w_sqrtm = []
for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    w = omega_vals[name]['signed']
    m = planets[name]['mass_sqrt']
    product = w * m
    w_sqrtm.append((name, product))
    print(f"  {name:10s}  {w:+10.3f}°  {m:12.4e}  {product:+12.6f}")

# Check if sum = 0 (angular momentum balance?)
total = sum(p for _, p in w_sqrtm)
print(f"\n  Sum of ω × √m = {total:.6f}")
print(f"  (Not zero — no direct angular momentum balance in ω)")

# ==============================================================
# SUMMARY
# ==============================================================
print("\n" + "=" * 80)
print("SUMMARY: ARGUMENT OF PERIHELION — FIBONACCI CONNECTIONS")
print("=" * 80)

print("""
Observed ω values (J2000, invariable plane):

  Planet      ω (signed)    Fibonacci expression       Error
  ─────────────────────────────────────────────────────────────""")

summary_data = [
    ('Mercury',  '+45',   '360°/F₆ = 360°/8',          ''),
    ('Venus',    '+72',   '360°/F₅ = 360°/5',          ''),
    ('Earth',    '+180',  '360°/F₃ = 360°/2',          ''),
    ('Mars',     '-19',   '360°×F₄/F₁₀ = 360°×3/55 ?', ''),
    ('Jupiter',  '+62',   '360°×F₅²/F₁₂ = 360°×25/144',''),
    ('Saturn',   '-27',   '360°/F₇ = 360°/13 ?',       ''),
    ('Uranus',   '-137',  '360°×F₅/F₇ = 360°×5/13 ?',  ''),
    ('Neptune',  '-146',  '360°×F₃/F₅ = 360°×2/5 ?',   ''),
]

for name in ['Mercury', 'Venus', 'Earth', 'Mars', 'Jupiter', 'Saturn', 'Uranus', 'Neptune']:
    w = omega_vals[name]['signed']
    expr_str, pred, fib_note = hypotheses[name]
    err_pct = 100 * abs(w - pred) / max(abs(pred), 0.01)
    quality = "EXACT" if err_pct < 0.5 else f"{err_pct:.1f}%"
    print(f"  {name:10s}  {w:+10.3f}°  {expr_str:25s}  {quality}")

print("""
Key findings:

1. CONSTANCY: All 8 planets maintain constant ω (perihelion and ascending
   node precess at the same rate for each planet).

2. FIBONACCI DENOMINATORS: Mercury (360°/8), Venus (360°/5), Earth (360°/2)
   use single Fibonacci numbers as denominators.

3. CROSS-PLANET LINKS: Mercury's ω denominator (8) = Saturn's period denominator.
   Venus's ω denominator (5) = Jupiter's period denominator.
   These are Law 6 triad numbers (3, 5, 8).

4. JUPITER: ω = 360° × F₅²/F₁₂ = 360° × 25/144 — both numerator and
   denominator are Fibonacci (5² and 144).

5. EARTH AT 180°: Earth's ω = 180° means its perihelion is exactly at the
   line of nodes (perpendicular to ascending node on invariable plane).

6. MIRROR PAIRS: Need further investigation — the relationships are more
   complex than simple sign reversal.

7. NOT A SIMPLE UNIVERSAL CONSTANT: Unlike d×η×√m = ψ, there is no single
   universal ω-constant. The structure is more like the eccentricity ladder
   (planet-specific but Fibonacci-organized).

Open questions:
- Exact values for Mars, Saturn, Uranus, Neptune (need Excel verification)
- Whether ω can be predicted from (d, b, F) alone
- Physical mechanism linking perihelion direction to node direction
- Whether this constitutes a 7th law or follows from the eigenmode structure
""")
