#!/usr/bin/env python3
"""
ECS analysis using boron-isotope CO₂ reconstructions instead of CenCO2PIP.

Motivation
----------
The §4.7 tightened analysis (climate_ecs_tight.py) showed that pre-MPT ice-share
contrasts are robust but absolute ECS values at iNHG-MPT and pre-iNHG are
biased HIGH because CenCO2PIP is 100-kyr-smoothed — short-period CO₂ amplitudes
are attenuated → CO₂ ΔF denominator is too small → ECS too high.

This script replaces CenCO2PIP with boron-isotope CO₂ reconstructions:

  - Chalk et al. 2017 (PNAS): MPT-focused, 4-1243 kyr, dt≈3 kyr (ODP 999)
  - Dyez et al. 2018 (P&P):   spans 4-4580 kyr, dt≈5 kyr (multi-site compilation)
  - Martinez-Boti et al. 2015 (Nature): 2338-3281 kyr, dt≈9 kyr (ODP 999)
  - de la Vega et al. 2020:   2019-4226 kyr, dt≈6.5 kyr (multi-site)

Boron δ¹¹B from planktonic foraminifera measures ancient seawater pH; combined
with [CO₃²⁻] estimates, gives atmospheric pCO₂ at sub-orbital resolution
without smoothing.

Strategy
--------
Use Dyez 2018 as the primary boron CO₂ proxy (covers all three regimes with
consistent calibration) for the per-regime analysis. Cross-validate with
Chalk 2017 (post-MPT high-resolution) and the others.

Expected outcomes
-----------------
  - iNHG-MPT ECS should drop from 6.4 K (CenCO2PIP-based) → consensus ~3-4 K
  - Pre-iNHG ECS should be definable (was 2.8-3.1 K but with smoothing-bias caveat)
  - Direct test of Martinez-Boti 2015 claim: "Pliocene ESS ≈ half Pleistocene"
"""

import json
import math
import sys
from pathlib import Path
import numpy as np
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

SCRIPTS_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPTS_DIR))
from milankovitch_climate_formula import (
    ClimateFormula, EIGHT_H, L1_LATTICE_INTEGERS,
    load_lr04, preprocess,
)

# ── Constants ────────────────────────────────────────────────────────────────
ALPHA = 5.35
F_2X = ALPHA * math.log(2)

GHG_MULT = 0.5
SL_PER_PERMIL = 120.0
F_PER_M_SL = 0.04
T_PER_PERMIL_LR04 = 2.5

N_BOOT = 200
BLOCK_KYR = 30
RNG_SEED = 20260606

SNYDER_FILE = Path("/home/dennis/code/3d/data/Snyder_Data_Figures/Source Data - Figure 1.xlsx")
SPRATT_FILE = Path("/home/dennis/code/3d/data/spratt-lisiecki-2016-sea-level.txt")
OUTPUT = Path("/home/dennis/code/3d/data/climate-ecs-boron.json")

BANDS = {
    "obliquity (35-50)":     [(35, 50)],
    "100-kyr band (75-130)": [(75, 130)],
    "precession (18-26)":    [(18, 26)],
}


# ── Loaders ──────────────────────────────────────────────────────────────────
def load_snyder():
    import openpyxl
    wb = openpyxl.load_workbook(SNYDER_FILE, data_only=True)
    ws = wb["1a-GAST reconstruction"]
    ages, t = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        if row[0] is None or row[2] is None: continue
        try:
            ages.append(float(row[0])); t.append(float(row[2]))
        except (TypeError, ValueError):
            continue
    wb.close()
    return np.asarray(ages), np.asarray(t)


def load_spratt():
    ages, sl = [], []
    with open(SPRATT_FILE) as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"): continue
            parts = s.split()
            if len(parts) < 6: continue
            try:
                a = float(parts[0]); v = parts[5]
                if v == "NaN": continue
                ages.append(a); sl.append(float(v))
            except ValueError:
                continue
    return np.asarray(ages), np.asarray(sl)


def load_boron_txt(path):
    """Generic loader for NOAA paleo trace-gas boron CO₂ format (tab-delimited).

    Format: first column is 'boron isotopes' (proxy name);
    col 5 = age_ka, col 8 = CO2_ppm.
    """
    ages, co2 = [], []
    with open(path, 'rb') as f:
        raw = f.read().decode('latin-1')
    for line in raw.split('\n'):
        parts = line.split('\t')
        if len(parts) < 8: continue
        if not parts[0].lower().startswith('boron'): continue
        try:
            a = float(parts[4])
            c = float(parts[7])
        except (ValueError, IndexError):
            continue
        ages.append(a); co2.append(c)
    return np.asarray(ages), np.asarray(co2)


def load_boron_xlsx(path, sheet='Boron isotopes'):
    """Generic loader for NOAA xlsx boron CO₂ format (same column convention)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    ages, co2 = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3: continue   # header rows
        try:
            a = float(row[4]); c = float(row[7])
        except (TypeError, ValueError, IndexError):
            continue
        ages.append(a); co2.append(c)
    wb.close()
    return np.asarray(ages), np.asarray(co2)


def interp_to_grid(ages_irregular, vals, grid_kyr):
    """Linear-interpolate irregular (age, val) onto a uniform grid.

    Sorts and dedupes by age first. Returns (grid_kyr, interp_vals) with vals
    set to nan outside the original age range.
    """
    order = np.argsort(ages_irregular)
    a = ages_irregular[order]
    v = vals[order]
    # Dedupe: if two points have same age, take mean
    uniq_a, uniq_idx = np.unique(a, return_inverse=True)
    uniq_v = np.zeros_like(uniq_a)
    counts = np.zeros_like(uniq_a)
    for k, val in zip(uniq_idx, v):
        uniq_v[k] += val
        counts[k] += 1
    uniq_v /= counts

    # Linear interp; values outside fall back to nearest endpoint
    interp = np.interp(grid_kyr, uniq_a, uniq_v,
                       left=float('nan'), right=float('nan'))
    return grid_kyr, interp


# ── Helpers ──────────────────────────────────────────────────────────────────
def fit_amps(t, y, regime):
    f = ClimateFormula()
    f.fit(t, y, regime=regime, normalize=True)
    std = f._fit_y_std
    return {n: {
        "amp": math.hypot(f._l1_a.get(n, 0.0), f._l1_b.get(n, 0.0)) * std,
        "phase": math.atan2(f._l1_b.get(n, 0.0), f._l1_a.get(n, 0.0)),
    } for n in L1_LATTICE_INTEGERS}


def block_idx(n, block_n, rng):
    if block_n >= n: block_n = n
    nb = (n + block_n - 1) // block_n
    starts = rng.integers(0, n - block_n + 1, size=nb)
    return np.concatenate([np.arange(s, s + block_n) for s in starts])[:n]


def boot_amps(t, y, regime, block_kyr, n_boot, rng):
    dt = float(np.median(np.diff(t)))
    block_n = max(1, int(round(block_kyr / dt)))
    out = {n: {"amp": [], "phase": []} for n in L1_LATTICE_INTEGERS}
    for _ in range(n_boot):
        idx = block_idx(len(t), block_n, rng)
        try:
            f = ClimateFormula()
            f.fit(t, y[idx], regime=regime, normalize=True)
            std = f._fit_y_std
            for n in L1_LATTICE_INTEGERS:
                a = f._l1_a.get(n, 0.0); b = f._l1_b.get(n, 0.0)
                out[n]["amp"].append(math.hypot(a, b) * std)
                out[n]["phase"].append(math.atan2(b, a))
        except Exception:
            continue
    return out


def wrap_pi(a):
    while a > math.pi: a -= 2 * math.pi
    while a <= -math.pi: a += 2 * math.pi
    return a


def pct(arr, q):
    return float(np.percentile(np.asarray(arr), q)) if len(arr) else float("nan")


def co2_forcing(d_co2, co2_ref):
    if co2_ref <= 0 or 1 + d_co2/co2_ref <= 0:
        return 0.0
    return ALPHA * math.log(1 + d_co2 / co2_ref)


def compute_freq_ice_fraction(t_lr, y_lr, t_sl, y_sl, win):
    tg_l, yg_l = preprocess(t_lr, y_lr, win, dt_kyr=1.0)
    tg_s, yg_s = preprocess(t_sl, y_sl, win, dt_kyr=1.0)
    amps_lr = fit_amps(tg_l, yg_l, win)
    amps_sl = fit_amps(tg_s, yg_s, win)
    f_ice = {}
    for n in L1_LATTICE_INTEGERS:
        a_lr = amps_lr[n]["amp"]
        a_sl = amps_sl[n]["amp"]
        if a_lr <= 1e-6:
            f_ice[n] = 0.5
            continue
        f_ice[n] = float(max(0.0, min(1.5, (a_sl / SL_PER_PERMIL) / a_lr)))
    return f_ice


# ── Per-regime analysis ──────────────────────────────────────────────────────
def analyze_regime(label, t_snyder, y_snyder, t_lr, y_lr,
                   co2_t_grid, co2_y_grid, co2_label,
                   win, f_ice_per_line, use_snyder_T, rng, caveat=""):
    # Preprocess T and LR04
    tg_l, yg_l = preprocess(t_lr, y_lr, win, dt_kyr=1.0)
    # CO₂: already on a regular grid (interpolated); restrict to win
    mask = (co2_t_grid >= win[0]) & (co2_t_grid <= win[1]) & ~np.isnan(co2_y_grid)
    if mask.sum() < 20:
        print(f"     ✗ {label}: insufficient CO₂ coverage in window {win}")
        return None, None
    tg_c = co2_t_grid[mask]
    yg_c = co2_y_grid[mask]
    co2_ref = float(np.nanmean(yg_c))

    if use_snyder_T:
        mask_s = (t_snyder >= win[0]) & (t_snyder <= win[1])
        if mask_s.sum() < 50:
            use_snyder_T = False
        else:
            sub_win = (max(win[0], float(t_snyder.min())),
                       min(win[1], float(t_snyder.max())))
            tg_s, yg_s = preprocess(t_snyder, y_snyder, sub_win, dt_kyr=1.0)

    # Baseline fits — CO₂ regime as tuple (custom window)
    base_L = fit_amps(tg_l, yg_l, win)
    base_C = fit_amps(tg_c, yg_c, (win[0], win[1]))
    if use_snyder_T:
        base_S = fit_amps(tg_s, yg_s, sub_win)
    else:
        base_S = {n: {"amp": base_L[n]["amp"] * T_PER_PERMIL_LR04,
                      "phase": base_L[n]["phase"] + math.pi}
                   for n in L1_LATTICE_INTEGERS}

    # Bootstrap
    boot_L = boot_amps(tg_l, yg_l, win, BLOCK_KYR, N_BOOT, rng)
    boot_C = boot_amps(tg_c, yg_c, (win[0], win[1]), BLOCK_KYR, N_BOOT, rng)
    if use_snyder_T:
        boot_S = boot_amps(tg_s, yg_s, sub_win, BLOCK_KYR, N_BOOT, rng)
    else:
        boot_S = {n: {
            "amp": [a * T_PER_PERMIL_LR04 for a in boot_L[n]["amp"]],
            "phase": [p + math.pi for p in boot_L[n]["phase"]],
        } for n in L1_LATTICE_INTEGERS}

    # Per-line ECS
    spectrum = []
    for n in L1_LATTICE_INTEGERS:
        aS = base_S[n]["amp"]; pS = base_S[n]["phase"]
        aL = base_L[n]["amp"]
        aC = base_C[n]["amp"]; pC = base_C[n]["phase"]
        if aS <= 0 or aL <= 0 or aC <= 0: continue
        sep = math.degrees(min(abs(wrap_pi(pC - pS)), math.pi))
        if sep > 90: continue

        f_ice = f_ice_per_line.get(n, 0.5)
        dF_co2 = co2_forcing(aC, co2_ref)
        dF_ghg = GHG_MULT * dF_co2
        dF_ice = aL * f_ice * SL_PER_PERMIL * F_PER_M_SL
        dF_tot = dF_co2 + dF_ghg + dF_ice
        if dF_tot < 1e-9: continue
        ecs = aS * F_2X / dF_tot

        boot_ecs = []
        for at, al, ac in zip(boot_S[n]["amp"], boot_L[n]["amp"], boot_C[n]["amp"]):
            if at <= 0 or al <= 0 or ac <= 0: continue
            dFc = co2_forcing(ac, co2_ref)
            dFi = al * f_ice * SL_PER_PERMIL * F_PER_M_SL
            dFt = dFc * (1 + GHG_MULT) + dFi
            if dFt < 1e-9: continue
            v = at * F_2X / dFt
            if 0 < v < 50: boot_ecs.append(v)

        spectrum.append({
            "n": n,
            "period_kyr": EIGHT_H / n,
            "amp_T_K": aS,
            "amp_LR04_permil": aL,
            "amp_CO2_ppm": aC,
            "f_ice": f_ice,
            "dF_co2": dF_co2,
            "dF_ice": dF_ice,
            "dF_total": dF_tot,
            "ice_share": dF_ice / dF_tot,
            "ecs_K": ecs,
            "ecs_p5_K": pct(boot_ecs, 5),
            "ecs_p95_K": pct(boot_ecs, 95),
        })

    # Per-band stats
    band_stats = {}
    for bname, ranges in BANDS.items():
        rows = [r for r in spectrum
                if any(lo <= r["period_kyr"] <= hi for lo, hi in ranges)]
        if not rows:
            band_stats[bname] = None
            continue
        w = np.array([r["amp_T_K"] for r in rows])
        ecs_v = np.array([r["ecs_K"] for r in rows])
        ice_v = np.array([r["ice_share"] for r in rows])
        ecs_p5_v = np.array([r["ecs_p5_K"] for r in rows])
        ecs_p95_v = np.array([r["ecs_p95_K"] for r in rows])
        band_stats[bname] = {
            "n_lines": len(rows),
            "ecs_weighted_K": float(np.average(ecs_v, weights=w)),
            "ecs_p5_K": float(np.average(ecs_p5_v, weights=w)),
            "ecs_p95_K": float(np.average(ecs_p95_v, weights=w)),
            "ice_share_weighted": float(np.average(ice_v, weights=w)),
            "amp_T_weighted_K": float(np.average(w, weights=w)),
            "amp_CO2_weighted_ppm": float(np.average(
                np.array([r["amp_CO2_ppm"] for r in rows]), weights=w)),
        }

    print(f"\n  ── {label}  (CO₂ source: {co2_label}) ──{caveat}")
    print(f"     CO₂_ref = {co2_ref:.1f} ppm  (mean over window)")
    print(f"     {'Band':<25}{'n':>3}{'ECS [5-95]':>22}{'ice%':>7}"
          f"{'ΔT (K)':>9}{'ΔCO₂ (ppm)':>12}")
    for bname, s in band_stats.items():
        if s:
            ci = f"[{s['ecs_p5_K']:.1f}-{s['ecs_p95_K']:.1f}]"
            print(f"     {bname:<25}{s['n_lines']:>3}"
                  f"  {s['ecs_weighted_K']:>5.2f} {ci:<14}"
                  f"{100*s['ice_share_weighted']:>6.0f}%"
                  f"{s['amp_T_weighted_K']:>9.3f}"
                  f"{s['amp_CO2_weighted_ppm']:>12.2f}")

    return spectrum, band_stats


def main():
    rng = np.random.default_rng(RNG_SEED)
    print("=" * 100)
    print("  ECS analysis with BORON-ISOTOPE CO₂ proxies (vs CenCO2PIP)")
    print(f"  Bootstrap N = {N_BOOT}, block = {BLOCK_KYR} kyr")
    print("=" * 100)

    # ── Load all proxies ──
    print("\n  Loading proxies ...")
    ages_s, t_s = load_snyder()
    t_lr, y_lr = load_lr04()
    t_sl, y_sl = load_spratt()

    # Boron datasets
    ages_chalk, co2_chalk = load_boron_txt("data/chalk-2017-boron-co2.txt")
    ages_dyez, co2_dyez = load_boron_txt("data/dyez-2018-boron-co2.txt")
    ages_mb, co2_mb = load_boron_txt("data/martinez-boti-2015-boron-co2.txt")
    ages_dlv, co2_dlv = load_boron_xlsx("data/delavega-2020-boron-co2.xlsx")

    print(f"    LR04:               {t_lr.min():.0f}-{t_lr.max():.0f} kyr (n={len(t_lr)})")
    print(f"    Snyder GAST:        {ages_s.min():.0f}-{ages_s.max():.0f} kyr (n={len(ages_s)})")
    print(f"    Spratt-Lisiecki:    {t_sl.min():.0f}-{t_sl.max():.0f} kyr (n={len(t_sl)})")
    print(f"    Chalk 2017:         {ages_chalk.min():.0f}-{ages_chalk.max():.0f} kyr (n={len(ages_chalk)}, dt≈3 kyr)")
    print(f"    Dyez 2018:          {ages_dyez.min():.0f}-{ages_dyez.max():.0f} kyr (n={len(ages_dyez)}, dt≈5 kyr)")
    print(f"    Martinez-Boti 2015: {ages_mb.min():.0f}-{ages_mb.max():.0f} kyr (n={len(ages_mb)}, dt≈9 kyr)")
    print(f"    de la Vega 2020:    {ages_dlv.min():.0f}-{ages_dlv.max():.0f} kyr (n={len(ages_dlv)}, dt≈7 kyr)")

    # ── Freq-dep ice fraction ──
    print("\n  Computing frequency-dependent ice fraction (Spratt-Lisiecki / LR04) ...")
    f_ice = compute_freq_ice_fraction(t_lr, y_lr, t_sl, y_sl, (0, 798))

    # ── Interpolate each boron record to 1-kyr grid ──
    print("\n  Interpolating boron records to 1-kyr grids ...")
    full_grid = np.arange(0, 5320, 1.0)
    _, co2_chalk_grid = interp_to_grid(ages_chalk, co2_chalk, full_grid)
    _, co2_dyez_grid  = interp_to_grid(ages_dyez,  co2_dyez,  full_grid)
    _, co2_mb_grid    = interp_to_grid(ages_mb,    co2_mb,    full_grid)
    _, co2_dlv_grid   = interp_to_grid(ages_dlv,   co2_dlv,   full_grid)
    print(f"    Chalk:        {(~np.isnan(co2_chalk_grid)).sum()} grid points covered")
    print(f"    Dyez:         {(~np.isnan(co2_dyez_grid)).sum()}")
    print(f"    Martinez-Boti:{(~np.isnan(co2_mb_grid)).sum()}")
    print(f"    de la Vega:   {(~np.isnan(co2_dlv_grid)).sum()}")

    # ── Per-regime analysis with boron CO₂ ──
    print("\n" + "=" * 100)
    print("  PER-REGIME ECS — boron-isotope CO₂ proxies")
    print("=" * 100)

    # Post-MPT: Chalk 2017 has 0-1243, restricted to 0-800
    spec_post_c, bands_post_c = analyze_regime(
        "Post-MPT (0-800 kyr, Chalk 2017 CO₂)",
        ages_s, t_s, t_lr, y_lr,
        full_grid, co2_chalk_grid, "Chalk 2017",
        (0, 800), f_ice, use_snyder_T=True, rng=rng,
    )

    # iNHG-MPT: Dyez 2018 covers it
    spec_inhg_d, bands_inhg_d = analyze_regime(
        "iNHG-MPT (1000-2000 kyr, Dyez 2018 CO₂)",
        ages_s, t_s, t_lr, y_lr,
        full_grid, co2_dyez_grid, "Dyez 2018",
        (1000, 2000), f_ice, use_snyder_T=True, rng=rng,
    )

    # Pre-iNHG: Dyez 2018 covers 2700-4580
    spec_pre_d, bands_pre_d = analyze_regime(
        "Pre-iNHG (2700-4580 kyr, Dyez 2018 CO₂, LR04 × κ)",
        ages_s, t_s, t_lr, y_lr,
        full_grid, co2_dyez_grid, "Dyez 2018",
        (2700, 4580), f_ice, use_snyder_T=False, rng=rng,
        caveat="  (LR04 × κ for T; Dyez 2018 extends to 4580 kyr)",
    )

    # Cross-check at pre-iNHG using de la Vega 2020 alone (truncated 2400-3200 sub-window)
    spec_pre_dv, bands_pre_dv = analyze_regime(
        "Pre-iNHG sub (2700-4200 kyr, de la Vega 2020 CO₂, LR04 × κ)",
        ages_s, t_s, t_lr, y_lr,
        full_grid, co2_dlv_grid, "de la Vega 2020",
        (2700, 4200), f_ice, use_snyder_T=False, rng=rng,
        caveat="  (cross-check vs Dyez 2018)",
    )

    # Cross-check at iNHG-boundary using Martinez-Boti
    spec_inhg_mb, bands_inhg_mb = analyze_regime(
        "iNHG→Pre boundary (2400-3200 kyr, Martinez-Boti 2015 CO₂, LR04 × κ)",
        ages_s, t_s, t_lr, y_lr,
        full_grid, co2_mb_grid, "Martinez-Boti 2015",
        (2400, 3200), f_ice, use_snyder_T=False, rng=rng,
        caveat="  (cross-check)",
    )

    # ── Cross-regime + cross-source comparison ──
    print("\n" + "=" * 100)
    print("  COMPARISON: boron vs CenCO2PIP per-regime ECS (post §4.7 tightened values)")
    print("=" * 100)
    cenco2pip_results = {
        # From data/climate-ecs-tight.json
        "post_mpt": {
            "obliquity (35-50)": {"ECS": 2.44, "ice": 65, "ΔCO2": "EPICA"},
            "100-kyr band (75-130)": {"ECS": 3.65, "ice": 59, "ΔCO2": "EPICA"},
            "precession (18-26)": {"ECS": 2.76, "ice": 64, "ΔCO2": "EPICA"},
        },
        "inhg_mpt_cenco2pip": {
            "obliquity (35-50)": {"ECS": 6.38, "ice": 95},
            "100-kyr band (75-130)": {"ECS": 5.04, "ice": 57},
            "precession (18-26)": {"ECS": 10.13, "ice": 95},
        },
        "pre_inhg_cenco2pip": {
            "obliquity (35-50)": {"ECS": 2.82, "ice": 95},
            "100-kyr band (75-130)": {"ECS": 2.91, "ice": 78},
            "precession (18-26)": {"ECS": 3.12, "ice": 92},
        },
    }

    print()
    print(f"  {'Regime/Band':<35}{'CenCO2PIP':>20}{'BORON':>20}{'Δ ECS':>14}")
    pairs = [
        ("post-MPT obliquity", "post_mpt", "obliquity (35-50)", bands_post_c),
        ("post-MPT 100-kyr",   "post_mpt", "100-kyr band (75-130)", bands_post_c),
        ("iNHG-MPT obliquity", "inhg_mpt_cenco2pip", "obliquity (35-50)", bands_inhg_d),
        ("iNHG-MPT 100-kyr",   "inhg_mpt_cenco2pip", "100-kyr band (75-130)", bands_inhg_d),
        ("pre-iNHG obliquity", "pre_inhg_cenco2pip", "obliquity (35-50)", bands_pre_d),
        ("pre-iNHG 100-kyr",   "pre_inhg_cenco2pip", "100-kyr band (75-130)", bands_pre_d),
    ]
    for label, ref_key, bname, bands_new in pairs:
        if bands_new is None or bands_new.get(bname) is None: continue
        ref = cenco2pip_results.get(ref_key, {}).get(bname)
        new_ecs = bands_new[bname]["ecs_weighted_K"]
        if ref:
            d = new_ecs - ref["ECS"]
            print(f"  {label:<35}{ref['ECS']:>18.2f} K{new_ecs:>18.2f} K"
                  f"{d:>+14.2f}")
        else:
            print(f"  {label:<35}{'—':>20}{new_ecs:>18.2f} K{'':>14}")

    # Output
    OUTPUT.write_text(json.dumps({
        "method": (
            "ECS analysis using boron-isotope CO₂ proxies instead of CenCO2PIP. "
            "Per-regime: post-MPT (Chalk 2017), iNHG-MPT (Dyez 2018), "
            "pre-iNHG (Dyez 2018 + de la Vega 2020 + Martinez-Boti 2015 cross-checks). "
            f"Bootstrap N={N_BOOT}, block={BLOCK_KYR} kyr."
        ),
        "constants": {
            "ghg_multiplier": GHG_MULT,
            "sl_per_permil_m": SL_PER_PERMIL,
            "forcing_per_m_sl_Wm2_per_m": F_PER_M_SL,
            "t_per_permil_lr04_K": T_PER_PERMIL_LR04,
        },
        "freq_dep_ice_fraction": {str(n): f_ice[n] for n in L1_LATTICE_INTEGERS},
        "regimes": {
            "post_mpt_chalk":       {"window_kyr": [0, 800],
                                     "by_band": bands_post_c, "spectrum": spec_post_c},
            "inhg_mpt_dyez":        {"window_kyr": [1000, 2000],
                                     "by_band": bands_inhg_d, "spectrum": spec_inhg_d},
            "pre_inhg_dyez":        {"window_kyr": [2700, 4580],
                                     "by_band": bands_pre_d, "spectrum": spec_pre_d},
            "pre_inhg_delavega":    {"window_kyr": [2700, 4200],
                                     "by_band": bands_pre_dv, "spectrum": spec_pre_dv},
            "inhg_boundary_mb":     {"window_kyr": [2400, 3200],
                                     "by_band": bands_inhg_mb, "spectrum": spec_inhg_mb},
        },
        "cenco2pip_reference": cenco2pip_results,
    }, indent=2, default=str))
    print(f"\n  Wrote: {OUTPUT}")


if __name__ == "__main__":
    main()
